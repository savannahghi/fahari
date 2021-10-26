from collections import OrderedDict
from typing import (
    Any,
    Dict,
    Generic,
    Iterator,
    List,
    Optional,
    Sequence,
    Tuple,
    Type,
    TypeVar,
    Union,
)

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.styles.named_styles import NamedStyle
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.fields import Field
from rest_framework.relations import ManyRelatedField, PrimaryKeyRelatedField
from rest_framework.serializers import Serializer

from fahari.common.serializers import AuditFieldsMixin

from .excel_io import ExcelIO, ProgressCallback
from .excel_template import ExcelIOTemplate

S = TypeVar("S", bound=Serializer)
S_Audit = TypeVar("S_Audit", bound=AuditFieldsMixin)
T = TypeVar("T", bound=ExcelIOTemplate)
DT = List[Dict[str, Any]]
_NFD = Dict[str, Union[Field, Dict[str, Any]]]  # Nested Fields Dict

_EXCEL_TYPES = (bool, int, float, str)


def flatten_fields(fields: _NFD, nested_entries_delimiter: str) -> Dict[str, Field]:
    """Flatten the nested fields in the given dict using the given delimiter.

    This is needed as `DRFSerializerExcelIOTemplate` doesn't works with nested data.
    """

    # Visit each of the dump_fields and extract the field names to be used as
    # column headers. If there are nested fields, flatten them.
    def visit(source_fields: _NFD, parent_key="") -> Dict[str, Field]:
        results: List[Tuple[str, Field]] = []
        for key, val in source_fields.items():
            new_key = parent_key + nested_entries_delimiter + key if parent_key else key
            if isinstance(val, dict):
                results.extend(visit(val, new_key).items())  # noqa
            else:
                results.append((new_key, val))
        return OrderedDict(results)

    return visit(fields)


class DRFSerializerExcelIO(Generic[S, T], ExcelIO[DT]):
    """`ExcelIO` implementation powered by DRF `Serializer`."""

    serializer_class: Optional[Type[S]] = None
    template_class: Optional[Type[T]] = None

    def __init__(
        self,
        serializer_class: Optional[Type[S]] = None,
        template_class: Optional[Type[T]] = None,
        dump_fields: Sequence[str] = None,
        nested_entries_delimiter="::",
        context: Optional[Dict[str, Any]] = None,
    ):
        super().__init__()
        self.serializer_class = serializer_class or self.serializer_class
        self.template_class = template_class or self.template_class
        self._dump_fields = dump_fields
        self._nested_entries_delimiter = nested_entries_delimiter
        self._context = context or {}

    @property
    def context(self) -> Dict[str, Any]:
        """Return the context of this excel io instance."""

        return self._context

    def dump_data(self, data: DT, progress_callback: ProgressCallback = None) -> Workbook:
        dump_fields = self._pick_dump_fields()
        nested_entries_delimiter = self.get_nested_entries_delimiter()
        template = self.get_template(fields=flatten_fields(dump_fields, nested_entries_delimiter))
        wb = Workbook()
        template.render(
            self._clean_dump_data(data, dump_fields),
            progress_callback=progress_callback,
            workbook=wb,
        )
        return wb

    def ingest_data(self, source: Workbook, progress_callback: ProgressCallback = None) -> DT:
        raise NotImplementedError("`ingest_data` must be implemented.")

    def generate_input_template(self, progress_callback: ProgressCallback = None) -> Workbook:
        raise NotImplementedError("`generate_input_template` must be implemented.")

    def get_fields(self) -> _NFD:
        """Return all the fields supported by this excel io instance."""

        # Visit each field and construct the final fields dict. For each
        # "serializer field", map it's fields as nested dict of fields.
        def visit(fields: Dict[str, Field]) -> _NFD:
            results = []
            for field_name, field in fields.items():
                if isinstance(field, Serializer):
                    results.append((field_name, visit(field.get_fields())))
                else:
                    results.append((field_name, field))
            return OrderedDict(results)

        return visit(self.get_serializer().get_fields())

    def get_nested_entries_delimiter(self) -> str:
        """Return the delimiter used to separate nested entries."""

        return self._nested_entries_delimiter

    def get_serializer(self, *args, **kwargs) -> S:
        """Return a serializer instance.

        The serializer instance will be used for validating and deserializing
        input, and for serializing output.
        """

        serializer_class = self.get_serializer_class()
        kwargs.setdefault("context", self.get_serializer_context())
        return serializer_class(*args, **kwargs)

    def get_serializer_class(self) -> Type[S]:
        """Return a class to use for the serializer. Defaults to using `self.serializer_class`."""

        assert self.serializer_class is not None, (
            "'%s' should either include a `serializer_class` attribute, "
            "or override the `get_serializer_class()` method." % self.__class__.__name__
        )
        return self.serializer_class

    def get_serializer_context(self) -> Dict[str, Any]:
        """Return extra context to be passed to a serializer instance."""

        context = {"excel_io": self}
        context.update(self._context)
        return context

    def get_template(self, *args, **kwargs) -> T:
        """Return a template instance.

        The template instance will be used for layout and formatting of excel
        workbooks.
        """

        template_class = self.get_template_class()
        kwargs.setdefault("serializer", self.get_serializer())
        if "fields" not in kwargs:  # This is expensive so only call it when needed
            kwargs["fields"] = flatten_fields(
                self.get_fields(), self.get_nested_entries_delimiter()
            )

        return template_class(*args, **kwargs)  # type: ignore

    def get_template_class(self) -> Type[T]:
        """Return a class to use for the template. Defaults to using `self.template_class`."""

        assert self.template_class is not None, (
            "'%s' should either include a `template_class` attribute, "
            "or override the `get_template_class()` method." % self.__class__.__name__
        )
        return self.template_class

    def _clean_dump_data(self, data: DT, dump_fields: _NFD) -> DT:
        nested_entries_delimiter = self.get_nested_entries_delimiter()

        # Visit each entry in the given data and denormalize/flatten it if it
        # is nested and also filter out unwanted/unrequested fields.
        def visit(entry: Dict[str, Any], fields: _NFD, parent_key="") -> Dict[str, Any]:
            results: List[Tuple[str, Any]] = []
            for key, val in fields.items():
                new_key = parent_key + nested_entries_delimiter + key if parent_key else key
                if isinstance(val, dict):
                    results.extend(visit(entry[key], val, new_key).items())  # noqa
                else:
                    results.append((new_key, entry[key]))

            return OrderedDict(results)

        return [visit(entry, dump_fields) for entry in data]

    def _pick_dump_fields(self) -> _NFD:
        """Return a dict consisting of only the selected dump fields.

        If no dump fields have been provided, then returns all the fields as is.
        """

        all_fields = self.get_fields()
        if not self._dump_fields:
            return all_fields

        nested_entry_delimiter = self.get_nested_entries_delimiter()

        # Visit each nested field and only pick/select the requested dump fields.
        def visit_horizontally(nested_fields: Sequence[str], fields: _NFD, results: _NFD) -> None:
            parent = nested_fields[0]
            if len(nested_fields) > 1:
                results.setdefault(parent, OrderedDict())
                visit_horizontally(nested_fields[1:], fields[parent], results[parent])
            else:
                results[parent] = fields[parent]

        # Visit all the fields and only pick/select the requested dump fields.
        def visit_vertically(dump_fields: Sequence[str], fields: _NFD) -> _NFD:
            results: _NFD = OrderedDict()
            for dump_field in dump_fields:
                nested_fields = dump_field.split(nested_entry_delimiter)
                parent = nested_fields[0]
                if len(nested_fields) > 1:
                    results.setdefault(parent, OrderedDict())
                    visit_horizontally(nested_fields[1:], fields[parent], results[parent])
                else:
                    results[parent] = fields[parent]

            return results

        return visit_vertically(self._dump_fields, all_fields)


class DRFSerializerExcelIOTemplate(Generic[S], ExcelIOTemplate[DT]):
    """A template that works with DRFSerializerExcelIO objects."""

    DATA_WORKSHEET_NAME = "Data"
    SCHEMA_WORKSHEET_NAME = "Schema"

    COLUMN_HEADER_BORDER_SIDE = Side(style="thick", color="00000000")
    REQUIRED_COLUMN_HEADER_STYLE_NAME = "Column Header(Required)"
    REQUIRED_COLUMN_HEADER_STYLE = NamedStyle(
        name=REQUIRED_COLUMN_HEADER_STYLE_NAME,
        font=Font(sz=14, b=True, color="00FF0000"),
        border=Border(
            left=COLUMN_HEADER_BORDER_SIDE,
            top=COLUMN_HEADER_BORDER_SIDE,
            right=COLUMN_HEADER_BORDER_SIDE,
            bottom=COLUMN_HEADER_BORDER_SIDE,
        ),
    )
    OPTIONAL_COLUMN_HEADER_STYLE_NAME = "Column Header(Optional)"
    OPTIONAL_COLUMN_HEADER_STYLE = NamedStyle(
        name=OPTIONAL_COLUMN_HEADER_STYLE_NAME,
        font=Font(sz=14, b=True, color="00008000"),
        border=Border(
            left=COLUMN_HEADER_BORDER_SIDE,
            top=COLUMN_HEADER_BORDER_SIDE,
            right=COLUMN_HEADER_BORDER_SIDE,
            bottom=COLUMN_HEADER_BORDER_SIDE,
        ),
    )

    def __init__(self, serializer: S, fields: Dict[str, Field]):
        super().__init__()
        self._serializer = serializer
        self._fields = fields

    def generate_input_template(
        self, workbook: Workbook, progress_callback: Optional[ProgressCallback] = None
    ) -> None:
        raise NotImplementedError("`generate_input_template` must be implemented.")

    def get_column_headers(self, for_input=False) -> Sequence[str]:
        """Return the headers for each of the columns in the final exported file."""

        if for_input:  # pragma: no branch
            return tuple(self.get_fields().keys())

        # If this is a dump, ignore write only, primary related and
        # many to many fields.
        fields = self.get_fields()
        return tuple(
            field_name
            for field_name, field in fields.items()
            if not (
                field.write_only or isinstance(field, (PrimaryKeyRelatedField, ManyRelatedField))
            )
        )

    def get_fields(self) -> Dict[str, Field]:
        return self._fields

    def get_input_fields(self) -> Sequence[Field]:
        raise NotImplementedError("`get_input_fields` must be implemented.")

    def get_readonly_fields(self) -> Sequence[Field]:
        raise NotImplementedError("`get_readonly_fields` must be implemented.")

    def get_required_fields(self) -> Sequence[Field]:
        raise NotImplementedError("`get_required_fields` must be implemented.")

    def get_serializer(self) -> S:
        return self._serializer

    def read(self, workbook: Workbook, progress_callback: Optional[ProgressCallback] = None) -> DT:
        raise NotImplementedError("`read` must be implemented.")

    def render(
        self, data: DT, workbook: Workbook, progress_callback: Optional[ProgressCallback] = None
    ) -> None:
        self._setup_workbook(workbook, False)
        self._setup_data_worksheet(workbook[self.DATA_WORKSHEET_NAME], False)
        self._dump_data(data, workbook[self.DATA_WORKSHEET_NAME])
        self._auto_size_columns(workbook[self.DATA_WORKSHEET_NAME])

    def _dump_data(self, data: DT, worksheet: Worksheet) -> None:
        fields = self.get_fields()
        for entry in data:
            worksheet.append(
                self._coax_to_excel_value(value)
                for value, field in zip(entry.values(), fields.values())
                if not (
                    # For a data dump, ignore write only, primary related
                    # and many to many fields.
                    field.write_only
                    or isinstance(field, (PrimaryKeyRelatedField, ManyRelatedField))
                )
            )

    def _setup_data_worksheet(self, worksheet: Worksheet, for_input=False) -> None:
        header_row = self.get_column_headers()

        worksheet.append(header_row)
        self._style_column_headers(
            worksheet, self.OPTIONAL_COLUMN_HEADER_STYLE_NAME, 0, len(header_row)
        )
        self._freeze_column_headers(worksheet, len(header_row))

    def _setup_workbook(self, workbook: Workbook, for_input=False) -> None:
        work_sheets = workbook.sheetnames  # Remove existing worksheets
        for sheet_name in work_sheets:
            del workbook[sheet_name]

        ws: Worksheet = workbook.create_sheet(title=self.DATA_WORKSHEET_NAME, index=0)  # noqa
        workbook.active = ws
        workbook.create_sheet(title=self.SCHEMA_WORKSHEET_NAME)

        workbook.add_named_style(self.REQUIRED_COLUMN_HEADER_STYLE)
        workbook.add_named_style(self.OPTIONAL_COLUMN_HEADER_STYLE)

    @staticmethod
    def _auto_size_columns(worksheet: Worksheet) -> None:
        # Currently in openpyxl, there is no way of setting a column's
        # optimal width as supported in MS Excel and LibraOffice Calc.
        # This implementation is a simple approximation of that feature.
        columns: Iterator[Sequence[Cell]] = worksheet.columns
        for column_cells in columns:
            cell_values_lengths = map(
                lambda cell: 0 if cell.value is None else len(str(cell.value)), column_cells
            )
            column_letter: str = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = (max(cell_values_lengths) * 1.2) + 5

    @staticmethod
    def _coax_to_excel_value(value: Any) -> Union[bool, float, int, str]:
        if type(value) in _EXCEL_TYPES:
            return value
        return "" if value is None else str(value)

    @staticmethod
    def _freeze_column_headers(
        worksheet: Worksheet, no_of_column_headers: int, headers_row_index=1
    ) -> None:
        constraint_cell_column = get_column_letter(no_of_column_headers + 1)
        constraint_cell_id = f"{constraint_cell_column}{headers_row_index + 1}"
        worksheet.freeze_panes = worksheet[constraint_cell_id]

    @staticmethod
    def _style_column_headers(
        worksheet: Worksheet,
        style_name: str,
        start_col_index: int,
        last_col_index: int,
        headers_row_index=1,
    ) -> None:
        for index in range(start_col_index, last_col_index):
            column: str = get_column_letter(index + 1)
            worksheet["{}{}".format(column, headers_row_index)].style = style_name


class AuditSerializerExcelIO(Generic[S_Audit, T], DRFSerializerExcelIO[S_Audit, T]):
    """ExcelIO implementation powered by DRF Serializer and that excludes audit fields.

    *NB: This will exclude all audit fields specified by the
    `get_audit_field_names` method including those of nested serializer
    fields.*
    """

    DEFAULT_AUDIT_FIELD_NAMES = (
        "active",
        "created",
        "created_by",
        "id",
        "updated",
        "updated_by",
        "url",
        "organisation",
    )

    def ingest_data(self, source: Workbook, progress_callback: ProgressCallback = None) -> DT:
        raise NotImplementedError("`ingest_data` must be implemented.")

    def generate_input_template(self, progress_callback: ProgressCallback = None) -> Workbook:
        raise NotImplementedError("`generate_input_template` must be implemented.")

    def get_audit_field_names(self) -> Sequence[str]:
        """Return the names of the audit field names to be excluded."""

        return self.DEFAULT_AUDIT_FIELD_NAMES

    def get_fields(self) -> _NFD:
        """Return all the fields supported by this excel io instance but exclude audit fields."""

        audit_fields_names = self.get_audit_field_names()

        # Visit each field and construct the final fields dict. For each
        # "serializer field", map it's fields as nested dict of fields.
        def visit(fields: Dict[str, Field]) -> _NFD:
            results = []
            for field_name, field in fields.items():
                if field_name in audit_fields_names:  # Skip audit fields
                    continue
                if isinstance(field, Serializer):
                    results.append((field_name, visit(field.get_fields())))
                else:
                    results.append((field_name, field))
            return OrderedDict(results)

        return visit(self.get_serializer().get_fields())
