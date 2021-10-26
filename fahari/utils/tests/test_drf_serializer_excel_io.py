from typing import Any, Callable

import pytest
from _pytest._code import ExceptionInfo
from django.test import TestCase
from model_bakery import baker
from openpyxl import Workbook

from fahari.common.models import Organisation
from fahari.common.serializers import FacilitySerializer
from fahari.ops.models import FacilitySystem, FacilitySystemTicket
from fahari.ops.serializers import FacilitySystemSerializer, FacilitySystemTicketSerializer
from fahari.utils.excel_utils import (
    AuditSerializerExcelIO,
    DRFSerializerExcelIO,
    DRFSerializerExcelIOTemplate,
)
from fahari.utils.excel_utils.drf_serializer_excel_io import flatten_fields


class AuditSerializerExcelIOTestCase(TestCase):
    """Tests for the `AuditSerializerExcelIO` class."""

    def setUp(self) -> None:
        self.excel_io = AuditSerializerExcelIO(
            serializer_class=FacilitySystemSerializer,
            template_class=DRFSerializerExcelIOTemplate,
        )
        self.organisation = baker.make(Organisation, organisation_name="Savannah Informatics")

    def test_get_fields(self) -> None:
        """Assert that the `get_fields` method doesn't return audit fields."""

        audit_field_names = self.excel_io.get_audit_field_names()
        excel_io_fields = self.excel_io.get_fields()
        for audit_field_name in audit_field_names:
            assert audit_field_name not in excel_io_fields

        # Ensure that even nested audit fields are excluded
        nested_fields = excel_io_fields["facility_data"]
        for audit_field_name in audit_field_names:
            assert audit_field_name not in nested_fields


class DRFSerializerExcelIOTestCase(TestCase):
    """Tests for the `DRFSerializerExcelIO` class."""

    def setUp(self) -> None:
        self.excel_io = DRFSerializerExcelIO(
            dump_fields=(
                "id",
                "facility_data:::name",
                "system_data:::name",
                "trainees",
                "version",
            ),
            nested_entries_delimiter=":::",
            serializer_class=FacilitySystemSerializer,
            template_class=DRFSerializerExcelIOTemplate,
        )
        self.organisation = baker.make(Organisation, organisation_name="Savannah Informatics")

    def test_correct_object_creation(self) -> None:  # noqa
        """Assert that object creation and initialization produces the expected object."""

        excel_io = DRFSerializerExcelIO(
            nested_entries_delimiter="[]:[]",
            serializer_class=FacilitySerializer,
            template_class=DRFSerializerExcelIOTemplate,
        )

        assert excel_io.context == {}
        assert excel_io.get_fields() is not None
        assert excel_io.get_nested_entries_delimiter() == "[]:[]"
        assert excel_io.get_serializer_class() is FacilitySerializer
        assert excel_io.get_serializer_context() == {"excel_io": excel_io}
        assert excel_io.get_template_class() is DRFSerializerExcelIOTemplate
        assert isinstance(excel_io.get_serializer(), FacilitySerializer)
        assert isinstance(excel_io.get_template(), DRFSerializerExcelIOTemplate)
        assert FacilitySerializer().get_fields().keys() <= excel_io.get_fields().keys()

    def test_dump_data(self) -> None:
        """Assert that writing to excel files works without any hitches expected."""

        tickets = baker.make(FacilitySystemTicket, 10, organisation=self.organisation)
        versions = baker.make(FacilitySystem, 10, organisation=self.organisation)
        tickets_data = FacilitySystemTicketSerializer(tickets, many=True).data
        versions_data = FacilitySystemSerializer(versions, many=True).data
        workbook = self.excel_io.dump_data(versions_data)
        workbook1 = DRFSerializerExcelIO(
            dump_fields=(
                "id",
                "facility_system_data::facility_data::name",
                "facility_system_data::system_data::name",
                "facility_system_data::trainees",
                "facility_system_data::version",
            ),
            serializer_class=FacilitySystemTicketSerializer,
            template_class=DRFSerializerExcelIOTemplate,
        ).dump_data(tickets_data)
        workbook2 = DRFSerializerExcelIO(
            serializer_class=FacilitySystemTicketSerializer,
            template_class=DRFSerializerExcelIOTemplate,
        ).dump_data(tickets_data)
        assert workbook is not None
        assert workbook1 is not None
        assert workbook2 is not None

    def test_ingest_data(self) -> None:
        """Assert that ingest data works as expected.

        This method is currently not implemented.
        """

        exc_info: ExceptionInfo[NotImplementedError]
        with pytest.raises(NotImplementedError) as exc_info:
            self.excel_io.ingest_data(Workbook())

        assert "`ingest_data` must be implemented." in exc_info.value.args

    def test_generate_input_template(self) -> None:
        """Assert that input template generation works as expected.

        This method is currently not implemented.
        """

        exc_info: ExceptionInfo[NotImplementedError]
        with pytest.raises(NotImplementedError) as exc_info:
            self.excel_io.generate_input_template()

        assert "`generate_input_template` must be implemented." in exc_info.value.args


class DRFSerializerExcelIOTemplateTestCase(TestCase):
    """Tests for the `DRFSerializerExcelIOTemplate` class."""

    def setUp(self) -> None:
        self.excel_io = DRFSerializerExcelIO(
            dump_fields=(
                "id",
                "facility_data:::name",
                "system_data:::name",
                "trainees",
                "version",
            ),
            nested_entries_delimiter=":::",
            serializer_class=FacilitySystemSerializer,
            template_class=DRFSerializerExcelIOTemplate,
        )
        self.excel_io_template = DRFSerializerExcelIOTemplate(
            fields=self.excel_io.get_fields(), serializer=self.excel_io.get_serializer()
        )
        self.organisation = baker.make(Organisation, organisation_name="Savannah Informatics")

    def test_correct_object_creation(self) -> None:
        """Assert that object creation and initialization produces the expected object."""

        delimiter = self.excel_io.get_nested_entries_delimiter()
        excel_io_template = DRFSerializerExcelIOTemplate(
            fields=flatten_fields(self.excel_io.get_fields(), delimiter),
            serializer=self.excel_io.get_serializer(),
        )

        assert excel_io_template.get_column_headers() is not None
        assert excel_io_template.get_fields() is not None
        assert excel_io_template.get_serializer() is not None

    def test_generate_input_template(self) -> None:
        """Assert that input template generation works as expected.

        This method is currently not implemented.
        """

        exc_info: ExceptionInfo[NotImplementedError]
        with pytest.raises(NotImplementedError) as exc_info:
            self.excel_io_template.generate_input_template(Workbook())

        assert "`generate_input_template` must be implemented." in exc_info.value.args

    def test_non_implemented_accessors(self) -> None:
        """Assert that are non implemented accessors raise `NotImplementedError`.

        *NB: This is hack and will be removed once all these accessors are
        implemented.*
        """

        accessors = ("get_input_fields", "get_readonly_fields", "get_required_fields")
        for accessor_name in accessors:
            accessor: Callable[[], Any] = getattr(self.excel_io_template, accessor_name)
            error_message = "`%s` must be implemented." % accessor_name
            with pytest.raises(NotImplementedError, match=error_message):
                accessor()

    def test_read(self) -> None:
        """Assert that reading data from an excel workbook works as expected."""

        exc_info: ExceptionInfo[NotImplementedError]
        with pytest.raises(NotImplementedError) as exc_info:
            self.excel_io_template.read(Workbook())

        assert "`read` must be implemented." in exc_info.value.args

    def test_render(self) -> None:
        """Assert that rendering of data in excel workbook works as expected."""

        versions = baker.make(FacilitySystem, 10, organisation=self.organisation)
        versions_data = FacilitySystemSerializer(versions, many=True).data
        # Even though this a DRFSerializerExcelIOTemplate test, an
        # "excel_io.dump_data" method(which will in turn call it's
        # "excel_io_template.render" method) has been used so that
        # any de-normalization of the given data can be performed
        # before rendering occurs.
        workbook = self.excel_io.dump_data(versions_data)

        assert workbook is not None
        assert len(workbook.sheetnames) == 2
        assert DRFSerializerExcelIOTemplate.DATA_WORKSHEET_NAME in workbook.sheetnames
        assert DRFSerializerExcelIOTemplate.SCHEMA_WORKSHEET_NAME in workbook.sheetnames
